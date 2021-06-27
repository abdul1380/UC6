'''
Created on 21 Jun 2020

@author: AbdulMannanRauf
'''
import os
import csv
import pandas as pd
import numpy as np

data_folder    =    'C:\\Users\\AbdulMannanRauf\\Desktop\\Notes\\Shakespeare'
data_file    =    'alllines.txt'

dir_2 = 'C:\\Users\\AbdulMannanRauf\\Desktop\\Notes\\Household' # used for os.path.join
dir_2_ = 'C:\\Users\\AbdulMannanRauf\\Desktop\\Notes\\Household\\' # used for pd.read ..   see the diff: with above
file_2_name = 'household_power_consumption.csv'
file_2 = os.path.join(dir_2,file_2_name)


if __name__ == '__main__':
    print("Chap 2")
    
#     f = open(os.path.join(data_folder,data_file),newline='')
#     print(f)
#     contents = f.read()
#     print(contents[:1000])
#     print(len(contents))

    with open(file_2,newline='') as csvfile,open(os.path.join(dir_2, 'temp.csv'),'w',newline='')  as  tempfile:
        csvreader  =  csv.reader(csvfile)
        csvwriter  =    csv.writer(tempfile)
        for row,  i in  zip(csvreader, range(10)):
                csvwriter.writerow(row)
    with  open(os.path.join(dir_2,    'temp.csv'),  newline='') as tempfile:
        csvreader   =   csv.reader(tempfile)
        for    row,    i    in    zip(csvreader,  range(3)): # here zip will let us parse only 2 line
            print(row)
    
    
    '''Working    with    CSV    files    with    the pandas    module    '''
            
    df    =    pd.read_csv(dir_2_+'temp.csv')
    #print(df.to_string()) # print enitre df
    #print(df.head(4))
    #print(df.iloc[:4, :5]) # prints 54 rows and 5 columns
    df.to_csv(dir_2_+'temp1.csv')
    
    '''Working    with    CSV    files    with    the  NumPy    module    '''
    arr    =    np.loadtxt(dir_2_+'temp.csv',  skiprows=1,  usecols=(2,3),  delimiter=',')
    print(arr)
    arr    =    np.genfromtxt(dir_2_+'temp.csv',    skip_header=1,    usecols=(2,3),    delimiter=',')
    print("\n",arr)
    
    
    
    """ Using    OpenPyXl    for    XLSX    files  
    OpenPyXl    is    a    Python    library    for    reading    and    writing    Excel    files."""
    
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    wb    =    Workbook()
    ws1    =    wb.active
    dest_filename = 'empty_book.xlsx'
    ws1.title = "range names"
    for row in range(1, 40):
        ws1.append(range(0,100,5)) # 40 rows added with values 0,5,10,15,20------,95
    ws2 = wb.create_sheet(title="Pi")
    ws2['F5'] = 2 * 3.14
    ws2.cell(column=1, row=5, value= 3.14)
    
    ws3 = wb.create_sheet(title="Data")
    for row in range(1, 20):
        for col in range(1, 15):
            _= ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
    print(ws3['A10'].value)
    wb.save(filename = dest_filename)
    
    """ The    following    code    reads    an    Excel    file,    manipulates    it,    and    saves    it"""
    
    df = pd.read_excel("empty_book.xlsx", sheet_name=0)  # this directory will be same as the one with the code
    print(df)
    df1=df*2
    df1.to_excel("empty_book_modified.xlsx")
    
    
        
""" Using    JSON    files    with    the    JSON module """  

    
import    json
with    open("sample_json.json")    as    json_file:
    for    line,i    in    zip(json_file,range(10)):
        print(line)
        json_data    =    json.loads(line)
        print(json_data)
        
        

    
    
    

    